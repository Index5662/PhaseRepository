
"""

import pandas as pd
import string

character_number_mapping = {string.ascii_lowercase[i]:i for i in range(len(string.ascii_lowercase))}

def col(cell_name: str):
    character = cell_name.split('-')[0]
    
    return character_number_mapping[character.lower()]

    
def row(cell_name: str):
    row_number = int(cell_name.split('-')[1])
    
    return row_number - 1


excel_file = pd.ExcelFile('Book1.xlsx', engine='openpyxl')

df = pd.DataFrame()
for sheet_name in excel_file.sheet_names:
    sheet_data = {'A': [], 'B': [], 'C': []}
    df_tmp = pd.read_excel(excel_file, sheet_name, header=None)
    sheet_data['A'].append(df_tmp.loc[row('A-2'), col('A-2')])
    sheet_data['B'].append(df_tmp.loc[row('B-1'), col('B-1')])
    sheet_data['C'].append(df_tmp.loc[row('C-2'), col('C-2')])
    sheet_data['A'].append(df_tmp.loc[row('A-1'), col('A-1')])
    sheet_data['B'].append(df_tmp.loc[row('B-2'), col('B-2')])
    sheet_data['C'].append(df_tmp.loc[row('C-1'), col('C-1')])
    df = pd.concat([df, pd.DataFrame(sheet_data)])

df.columns = ['A', 'B', 'C']
    
df.head()
"""


"""
import pandas as pd

dataframe1 = pd.read_excel('Book1.xlsx')
print(dataframe1)
"""

"""
import openpyxl
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Book1.xlsx")
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)
"""



# Python3 code to select
# data from excel
import xlwings as xw
 
# Specifying a sheet
ws = xw.Book("Book1.xlsx").sheets['Sheet1']
 
# Selecting data from
# a single cell
v1 = ws.range("A1:A7").value
v2 = ws.range("F5").value

print("Result:", v1, v2)

#works but need to find way to automate for all rows 



"""
#xx1
#Not Working

import openpyxl
 
dataframe = openpyxl.load_workbook("Book1.xlsx")
dataframe1 = dataframe.active
 
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)


"""

"""
output for xx1

Result: ['Name  Age    Stream  Percentage', 
'0      Ankit   18      Math          95', 
'1      Rahul   19   Science          90', 
'2    Shaurya   20  Commerce          85', 
'3  Aishwarya   18      Math          80', 
'4   Priyanka   19   Science          75', 
None]
        
"""

"""
#this workes but it is not the right system of order for the files ext

import os
import re

# from ez_setup import use_setuptools
# use_setuptools()

from setuptools import setup

MAIN_PKG = 'xlsx'

thisdir = os.path.dirname(__file__)

# history_path = os.path.join(thisdir, 'HISTORY.rst')
init_py_path = os.path.join(thisdir, MAIN_PKG, '__init__.py')
license_path = os.path.join(thisdir, 'LICENSE')
readme_path = os.path.join(thisdir, 'README.rst')

# with open(history_path) as f:
#     history = f.read()
with open(license_path) as f:
    license = f.read()
with open(readme_path) as f:
    readme = f.read()
with open(init_py_path) as f:
    version = re.search("__version__ = '([^']+)'", f.read()).group(1)

NAME = 'python-xlsx'
VERSION = version
DESCRIPTION = (
    'Create and modify Excel .xlsx files'
)

"""

"""
import xlwings as xw
ws = xw.Book("Book1.xlsx").sheets['Sheet1']

Loopnumber = 1

while i <= LoopNumber:                #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)


 v1 = ws.range("A1:A7").value
 v2 = ws.range("F5").value

print("Result:", v1, v2)

#works but need to find way to automate for all rows 


import requests
from bs4 import BeautifulSoup
from csv import writer
import uuid
import time
from datetime import datetime
import xlwings as xw
"""



RowNumber = 10

i=1
ws = xw.Book("Book1.xlsx").sheets['Sheet1']

while i <= RowNumber:                  #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)
 
 i = str(i)

 var1 = "A" + str(i)
 v2 = ws.range(var1).value

 i = int(i)

while i <= RowNumber:                  #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)
 
 i = str(i)
 
 var1 = "B" + str(i)
 v2 = ws.range(var1).value

 i = int(i)

while i <= RowNumber:                  #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)
 
 i = str(i)
 
 var1 = "C" + str(i)
 v2 = ws.range(var1).value

 i = int(i)

while i <= RowNumber:                  #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)
 
 i = str(i)
 
 var1 = "D" + str(i)
 v2 = ws.range(var1).value

 i = int(i)

while i <= RowNumber:                  #change 0 to number of loops you want +1 and pages to go through
 i = i + 1   
 print(i)
 
 i = str(i)
 
 var1 = "E" + str(i)
 v2 = ws.range(var1).value

 i = int(i)

#A+B
#B+D

dataframe = openpyxl.load_workbook("Book1.xlsx")
dataframe1 = dataframe.active
 
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)

with open(history_path) as f:
#     history = f.read()
with open(license_path) as f:
    license = f.read()
with open(readme_path) as f:
    readme = f.read()
with open(init_py_path) as f:
    version = re.search("__version__ = '([^']+)'", f.read()).group(1)
     
